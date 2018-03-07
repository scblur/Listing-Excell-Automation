from tkinter import Tk
from tkinter.filedialog import askopenfilename
import openpyxl

# Open file and get the file path
def get_file_path():
    Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
    file_path = askopenfilename() # show an "Open" dialog box and return the path to the selected file
    return file_path

# Function call
path = str(get_file_path())
print(path)

# Load Workbook
wb = openpyxl.load_workbook(path)

print(wb.get_sheet_names())

# Get the Sheet in an Object
sheet = wb.get_sheet_by_name("Sheet1")
print(sheet.title)

# Get the Max Rows and Max Columns in Sheet1
row_count = sheet.max_row
column_count = sheet.max_column
print('row_count{} column_count{}'.format(row_count,column_count))

sheet.title = "New Title"

# Create new Sheet
wb.create_sheet(title='NewSheet') # At the end
wb.create_sheet(title="at index 2",index=2) # At index 2

# Delete a Sheet
std = wb.get_sheet_by_name("NewSheet")
wb.remove_sheet(std)

print(wb.get_sheet_names())

# Saving the new Workbook in the same directory
before_sheetname = path.split('/')[-1]
print(before_sheetname)
new_sheetname = 'basefile2'
new_sheetname = new_sheetname + '.xlsx'
after_sheet_path = path.replace(before_sheetname, new_sheetname)
print(after_sheet_path)

# wb.save(path+new_sheet_name)
wb.save(after_sheet_path) # Save as
