import openpyxl
from tkinter import Tk
from tkinter.filedialog import askopenfilename

def get_file_path():
    Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
    file_path = askopenfilename() # show an "Open" dialog box and return the path to the selected file
    return file_path

path1 = str(get_file_path())
path2 = str(get_file_path())

wb = openpyxl.load_workbook(path1)
sheet = wb['Sheet1']

wb2 = openpyxl.load_workbook(path2)
sheet2 = wb2['Sheet2']

rows = sheet.max_row
columns = sheet.max_column

# Defining a 2D list to hold Data
listab = []

# Creating 2D list
for i in range(1,rows+1):
    listab.append([])

# Get all the data into a list
for r in range(1,rows+1):
    for c in range(1,columns+1):
        e = sheet.cell(row=r,column=c)
        listab[r-1].append(e.value)

print(listab)

# Writing the data to Sheet2
for r in range(1,rows+1):
    for c in range(1,columns+1):
        j = sheet2.cell(row=r,column=c)
        j.value = listab[r-1][c-1]

wb2.save(path2)
